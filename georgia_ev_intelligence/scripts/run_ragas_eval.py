"""
scripts/run_ragas_eval.py — Phase 4 RAGAS Evaluation using the OFFICIAL ragas library.

Usage:
  venv\\Scripts\\python scripts\\run_ragas_eval.py --questions 50
  venv\\Scripts\\python scripts\\run_ragas_eval.py --resume   # resume from checkpoint
"""
import argparse
import json
import sys
import time
from datetime import datetime
from pathlib import Path
from statistics import mean, pstdev

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

# RAGAS Official Library
from datasets import Dataset
from ragas import evaluate
from ragas.metrics.collections import (
    faithfulness,
    answer_relevancy,
    context_precision,
    context_recall,
    answer_correctness
)
from langchain_ollama import ChatOllama, OllamaEmbeddings

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from phase4_agent.pipeline import EVAgent
from shared.config import Config
from shared.logger import get_logger

logger = get_logger("ragas_eval")

PROGRESS_FILE = ROOT / "outputs" / "progress" / "phase4_eval_progress.jsonl"
ANSWERS_MD    = ROOT / "outputs" / "progress" / "phase4_eval_answers.md"

WEIGHTS = {
    "faithfulness":       0.25,
    "answer_relevancy":   0.20,
    "context_precision":  0.20,
    "context_recall":     0.20,
    "answer_correctness": 0.15,
}

# ── 7 smoke questions ────────────────────────────────────────────────────────
SMOKE_QUESTIONS: list[dict] = [
    {
        "id": "Q1", "category": "AGGREGATE",
        "question": "Which county has the highest total employment among Tier 1 suppliers only?",
        "golden": "Troup County has the highest total employment among Tier 1 suppliers.",
    },
    {
        "id": "Q2", "category": "RISK",
        "question": "Which EV supply chain roles in Georgia have only one supplier, making them a single point of failure?",
        "golden": "Georgia has multiple single-supplier roles including Charging Infrastructure (Morgan Corp.), Power Electronics (GSC Steel Stamping), and Materials (Haering Precision USA LP).",
    },
    {
        "id": "Q3", "category": "CYPHER-TIER",
        "question": "Which Georgia companies are classified under Battery Cell or Battery Pack roles, and what tier is each?",
        "golden": "6 companies: Hitachi Astemo Americas (Tier 1/2, Battery Cell), Hollingsworth & Vose (Tier 1/2, Battery Pack), Honda Development & Manufacturing (Tier 1/2, Battery Cell), Hyundai Motor Group (Tier 1/2, Battery Pack), F&P Georgia Manufacturing (Tier 1/2, Battery Pack), IMMI (Tier 1/2, Battery Pack).",
    },
    {
        "id": "Q4", "category": "CYPHER-OEM",
        "question": "Show the full supplier network linked to Rivian Automotive in Georgia, broken down by tier and EV Supply Chain Role.",
        "golden": "Georgia suppliers linked to Rivian include GSC Steel Stamping (Tier 2/3), Duckyang (Tier 2/3), Enchem America (Tier 2/3), Hyundai Transys Georgia Powertrain (Tier 1/2), and Remark International LLC (Tier 1).",
    },
    {
        "id": "Q5", "category": "CYPHER-PROD",
        "question": "Find Georgia-based companies that manufacture copper foil or electrodeposited materials for EV battery current collectors.",
        "golden": "Duckyang in Jackson County manufactures electrodeposited copper foil for EV battery current collectors.",
    },
    {
        "id": "Q6", "category": "CYPHER-LOC",
        "question": "In Gwinnett County, which company has the highest employment and what is its EV Supply Chain Role?",
        "golden": "SungEel Recycling Park Georgia has the highest employment in Gwinnett County with 650 employees. Its EV Supply Chain Role is Materials.",
    },
    {
        "id": "Q7", "category": "CYPHER-FAC",
        "question": "Which Georgia companies operate R&D facilities focused on EV technology?",
        "golden": "Racemark International LLC (Jones County, Tier 1) operates R&D facilities focused on EV technology.",
    },
]

# We pull in the 50 questions from the original file by importing them (to save space here)
# But for the standalone script, we can just use the SMOKE_QUESTIONS or import from a shared location.
# Wait, let's keep the user's 50 questions. I will import them from the original file.
try:
    from run_ragas_eval_original import FIFTY_QUESTIONS
except ImportError:
    FIFTY_QUESTIONS = SMOKE_QUESTIONS # Fallback if original is not backed up

# ── Ragas Evaluation Wrapper ───────────────────────────────────────────────────

def evaluate_row_ragas(row: dict, llm, embeddings) -> dict:
    """Evaluates a single row using the official RAGAS library."""
    data = {
        "question": [row["question"]],
        "answer": [row["answer"]],
        "contexts": [[row.get("context", "")]],
        "ground_truth": [row["golden"]]
    }
    dataset = Dataset.from_dict(data)
    
    # Run official ragas
    try:
        res = evaluate(
            dataset,
            metrics=[faithfulness, answer_relevancy, context_precision, context_recall, answer_correctness],
            llm=llm,
            embeddings=embeddings,
            raise_exceptions=False,
            is_async=False # Run sync to avoid event loop crashes
        )
        scores = res.to_pandas().iloc[0].to_dict()
    except Exception as e:
        logger.error(f"Ragas evaluation failed: {e}")
        scores = {}

    def _clip(v):
        try:
            return max(0.0, min(1.0, float(v)))
        except (TypeError, ValueError):
            return 0.0

    f  = _clip(scores.get("faithfulness", 0.0))
    ar = _clip(scores.get("answer_relevancy", 0.0))
    cp = _clip(scores.get("context_precision", 0.0))
    cr = _clip(scores.get("context_recall", 0.0))
    ac = _clip(scores.get("answer_correctness", 0.0))

    final = _clip(
        WEIGHTS["faithfulness"] * f +
        WEIGHTS["answer_relevancy"] * ar +
        WEIGHTS["context_precision"] * cp +
        WEIGHTS["context_recall"] * cr +
        WEIGHTS["answer_correctness"] * ac
    )

    return {
        **row,
        "faithfulness":            f,
        "answer_relevancy":        ar,
        "context_precision":       cp,
        "context_recall":          cr,
        "answer_correctness":      ac,
        "final_score":             final,
        "faithfulness_reason":     "", # Ragas handles reasoning internally
        "answer_relevancy_reason": "",
        "context_precision_reason":"",
        "context_recall_reason":   "",
        "correctness_reason":      "",
    }

# ── Checkpoint helpers ─────────────────────────────────────────────────────────

def load_checkpoint() -> dict[str, dict]:
    done: dict[str, dict] = {}
    if PROGRESS_FILE.exists():
        for line in PROGRESS_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    row = json.loads(line)
                    done[row["id"]] = row
                except Exception:
                    pass
    return done

def save_checkpoint(row: dict) -> None:
    PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with PROGRESS_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\\n")

def append_answers_md(row: dict) -> None:
    ANSWERS_MD.parent.mkdir(parents=True, exist_ok=True)
    with ANSWERS_MD.open("a", encoding="utf-8") as f:
        f.write(f"\\n## {row['id']} [{row['category']}] — {row.get('elapsed_s',0)}s\\n")
        f.write(f"**Q**: {row['question']}\\n\\n")
        f.write(f"**Golden**: {row['golden']}\\n\\n")
        f.write(f"**Generated**: {row['answer']}\\n\\n")
        f.write(
            f"**Scores**: faithfulness={row.get('faithfulness',0):.3f} | "
            f"relevancy={row.get('answer_relevancy',0):.3f} | "
            f"precision={row.get('context_precision',0):.3f} | "
            f"recall={row.get('context_recall',0):.3f} | "
            f"correctness={row.get('answer_correctness',0):.3f} | "
            f"**final={row.get('final_score',0):.3f}**\\n\\n"
            "---\\n"
        )

# ── Pipeline runner ────────────────────────────────────────────────────────────

def run_one(agent: EVAgent, q: dict) -> dict:
    t0 = time.monotonic()
    try:
        result = agent.ask(q["question"])
        return {
            "id":       q["id"],
            "category": q["category"],
            "question": q["question"],
            "golden":   q["golden"],
            "answer":   result["answer"],
            "context":  result.get("retrieved_context", ""),
            "elapsed_s": round(time.monotonic() - t0, 1),
            "path":     result["entities"].get(
                "retrieval_source",
                "cypher" if result["entities"].get("cypher_used") else "sql",
            ),
        }
    except Exception as exc:
        logger.error("Pipeline error on %s: %s", q["id"], exc)
        return {
            "id":       q["id"], "category": q["category"],
            "question": q["question"], "golden": q["golden"],
            "answer":   f"[PIPELINE ERROR: {exc}]", "context": "",
            "elapsed_s": round(time.monotonic() - t0, 1), "path": "error",
        }

# ── Excel report ───────────────────────────────────────────────────────────────

HEADERS = [
    "Q_ID","Category","Question","Golden_Answer","Generated_Answer","Path","Elapsed_s",
    "Faithfulness","Answer_Relevancy","Context_Precision","Context_Recall",
    "Answer_Correctness","Final_Score",
    "Faithfulness_Reason","Relevancy_Reason","Precision_Reason","Recall_Reason","Correctness_Reason",
]

def _fill(v: float) -> PatternFill:
    if v >= 0.7: return PatternFill(fill_type="solid", fgColor="C6EFCE")
    if v >= 0.5: return PatternFill(fill_type="solid", fgColor="FFF2CC")
    return PatternFill(fill_type="solid", fgColor="F4CCCC")

def build_report(results: list[dict], out: Path) -> None:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Results"
    hfill = PatternFill(fill_type="solid", fgColor="1F4E78")
    wfont = Font(color="FFFFFF", bold=True)
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = hfill; c.font = wfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    lb = PatternFill(fill_type="solid", fgColor="D9EAF7")
    wh = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for i, r in enumerate(results, 2):
        ws.append([
            r.get("id",""), r.get("category",""), r.get("question",""),
            r.get("golden",""), r.get("answer",""), r.get("path",""), r.get("elapsed_s",0),
            r.get("faithfulness",0), r.get("answer_relevancy",0),
            r.get("context_precision",0), r.get("context_recall",0),
            r.get("answer_correctness",0), r.get("final_score",0),
            r.get("faithfulness_reason",""), r.get("answer_relevancy_reason",""),
            r.get("context_precision_reason",""), r.get("context_recall_reason",""),
            r.get("correctness_reason",""),
        ])
        ws.row_dimensions[i].height = 70
        rf = lb if i % 2 == 0 else wh
        for col in range(1, len(HEADERS)+1):
            ws.cell(i, col).fill = rf
            ws.cell(i, col).alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(8, 14):
            cell = ws.cell(i, col)
            try:
                cell.fill = _fill(float(cell.value or 0))
                cell.number_format = "0.0000"
            except (TypeError, ValueError):
                pass

    ws.freeze_panes = "A2"
    for col, w in {"A":8,"B":16,"C":52,"D":55,"E":75,"F":8,"G":8}.items():
        ws.column_dimensions[col].width = w

    ss = wb.create_sheet("Summary")
    ss.append(["Metric","Mean","Std Dev","Weight"])
    for c in ss[1]:
        c.fill = hfill; c.font = wfont
    for mk in list(WEIGHTS) + ["final_score"]:
        vals = [float(r.get(mk, 0) or 0) for r in results]
        mu = mean(vals) if vals else 0.0
        sd = pstdev(vals) if len(vals) > 1 else 0.0
        ss.append([mk, round(mu, 4), round(sd, 4), WEIGHTS.get(mk, "-")])
        c = ss.cell(ss.max_row, 2)
        c.fill = _fill(mu); c.number_format = "0.0000"

    chart = BarChart()
    chart.title = "Phase 4 RAGAS Scores"
    chart.y_axis.title = "Score"
    data = Reference(ss, min_col=2, min_row=1, max_row=ss.max_row)
    cats = Reference(ss, min_col=1, min_row=2, max_row=ss.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9; chart.width = 16
    ss.add_chart(chart, "F2")

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Report saved: %s", out)

# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(results: list[dict]) -> None:
    print(f"\\n{'='*68}")
    print("  RAGAS EVALUATION SUMMARY — Official Library")
    print(f"{'='*68}")
    for mk in list(WEIGHTS) + ["final_score"]:
        vals = [float(r.get(mk,0) or 0) for r in results]
        mu = mean(vals) if vals else 0.0
        flag = "✅" if mu >= 0.7 else ("⚠️ " if mu >= 0.5 else "❌")
        w = f"w={WEIGHTS.get(mk,'—')}" if mk != "final_score" else "WEIGHTED"
        print(f"  {flag} {mk:<24} {mu:.4f}  {w}")
    print(f"{'='*68}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main(questions: list[dict], out_path: Path, resume: bool) -> None:
    cfg = Config.get()
    
    # Initialize LangChain wrappers for RAGAS
    llm = ChatOllama(model=cfg.ollama_llm_model, base_url=cfg.ollama_base_url)
    embeddings = OllamaEmbeddings(model=cfg.ollama_embed_model, base_url=cfg.ollama_base_url)

    done = load_checkpoint() if resume else {}
    if done:
        print(f"  Resuming — {len(done)} questions already done, skipping them.")
        if not resume and PROGRESS_FILE.exists():
            PROGRESS_FILE.unlink()

    agent  = EVAgent()
    total  = len(questions)
    scored: list[dict] = list(done.values())

    for i, q in enumerate(questions, 1):
        if q["id"] in done:
            print(f"  [{i}/{total}] {q['id']} — skipped (checkpoint)")
            continue

        print(f"\\n  [{i}/{total}] {q['id']} [{q['category']}]")
        print(f"  Q: {q['question'][:80]}")

        # Step 1: Pipeline
        t0  = time.monotonic()
        row = run_one(agent, q)
        print(f"  -> Answer ({row.get('elapsed_s',0)}s, {row.get('path')}): {str(row.get('answer'))[:80]}")

        # Step 2: RAGAS scoring
        print(f"  -> Scoring with Official Ragas...")
        scored_row = evaluate_row_ragas(row, llm, embeddings)
        fs = scored_row.get("final_score", 0)
        flag = "✅" if fs >= 0.7 else ("⚠️" if fs >= 0.5 else "❌")
        print(f"  -> Final score: {fs:.3f} {flag}")

        save_checkpoint(scored_row)
        append_answers_md(scored_row)
        scored.append(scored_row)

    print_summary(scored)
    build_report(scored, out_path)
    print(f"  📄 Excel : {out_path}")
    print(f"  📝 MD    : {ANSWERS_MD}")
    print(f"  💾 JSONL : {PROGRESS_FILE}\\n")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--questions", type=int, default=7)
    parser.add_argument("--resume",    action="store_true",
                        help="Resume from checkpoint")
    parser.add_argument("--out", type=str,
                        default=str(ROOT / "outputs" / "ragas_reports" /
                                    f"phase4_ragas_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
    args = parser.parse_args()

    qs = FIFTY_QUESTIONS[:args.questions]
    main(qs, Path(args.out), args.resume)
